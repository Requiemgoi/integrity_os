# -*- coding: utf-8 -*-
"""
Import ILI (In-Line Inspection) data from Excel files.
Supports the standard format from Zaklyuchitelny_Excel.xlsx
"""
import os
import uuid
from datetime import datetime
from typing import Optional, List, Dict, Any

import openpyxl
from sqlalchemy.orm import Session

from ..models import Defect


def calculate_severity(max_depth_percent: Optional[float], erf_b31g: Optional[float] = None) -> str:
    """
    Calculate defect severity based on depth and ERF.
    ERF < 1.0 means the defect is critical (failure expected).
    """
    if erf_b31g is not None and erf_b31g < 1.0:
        return "critical"
    if erf_b31g is not None and erf_b31g < 1.25:
        return "high"
    
    if max_depth_percent is None:
        return "low"
    if max_depth_percent >= 80:
        return "critical"
    elif max_depth_percent >= 50:
        return "high"
    elif max_depth_percent >= 20:
        return "medium"
    return "low"



def parse_float(value) -> Optional[float]:
    """Safely convert value to float, returning None for empty/whitespace strings."""
    if value is None:
        return None
    # Handle numeric types
    if isinstance(value, (int, float)):
        # Check for NaN
        if isinstance(value, float) and (value != value):  # NaN check
            return None
        return float(value)
    # Handle strings
    if isinstance(value, str):
        value = value.strip()
        # Return None for empty strings, whitespace, or just spaces
        if not value or value == '' or value.isspace():
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None
    # Try to convert other types
    try:
        # Handle cases where value might be a cell object from openpyxl
        if hasattr(value, 'value'):
            return parse_float(value.value)
        result = float(value)
        # Check for NaN
        if result != result:  # NaN check
            return None
        return result
    except (ValueError, TypeError, AttributeError):
        return None


def parse_int(value) -> Optional[int]:
    """Safely convert value to int, returning None for empty/whitespace strings."""
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if not (value != value) else None  # Check for NaN
    if isinstance(value, str):
        value = value.strip()
        if not value or value == '':
            return None
        try:
            return int(float(value))  # Convert via float to handle "1.0" -> 1
        except (ValueError, TypeError):
            return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None



def parse_string(value) -> Optional[str]:
    """Safely convert value to string, returning None for empty/whitespace strings."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return str(value) if value else None


def parse_orientation(value) -> Optional[str]:
    """Convert datetime.time orientation to string format."""
    if value is None:
        return None
    if hasattr(value, 'strftime'):
        return value.strftime("%H:%M")
    return str(value)


def import_anomalies_from_excel(
    db: Session,
    file_path: str,
    pipeline_code: str,
    inspection_date: datetime = None,
    sheet_name: str = "Аномалии"
) -> Dict[str, Any]:
    """
    Import anomalies from ILI Excel report.
    
    Args:
        db: Database session
        file_path: Path to Excel file
        pipeline_code: Pipeline code (MT-01, MT-02, MT-03)
        inspection_date: Date of inspection (defaults to now)
        sheet_name: Name of the sheet with anomalies
        
    Returns:
        Dict with import statistics
    """
    if inspection_date is None:
        inspection_date = datetime.now()
    
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    
    if sheet_name not in wb.sheetnames:
        return {"error": f"Sheet '{sheet_name}' not found", "imported": 0}
    
    ws = wb[sheet_name]
    
    # Find header row (row with "шов на [м]")
    header_row = None
    for row_num in range(1, 20):
        row = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
        if row[0] and "шов" in str(row[0]).lower():
            header_row = row_num
            break
    
    if header_row is None:
        return {"error": "Header row not found", "imported": 0}
    
    # Column mapping (0-indexed)
    COL_MAP = {
        "weld_distance": 0,      # шов на [м]
        "section_number": 1,     # № секции
        "section_length": 2,     # длина секции [м]
        "wall_thickness": 3,     # прив.ТС [мм]
        "measured_distance": 5,  # измер. расст. [м]
        "orientation": 6,        # ориентация
        "defect_type": 7,        # тип аномалии
        "identification": 8,     # идентификация
        "comment": 9,            # комментарий
        "external_size": 10,     # внеш. размеры
        "length_mm": 11,         # длина [мм]
        "width_mm": 12,          # ширина [мм]
        "max_depth_percent": 13, # макс. глубина [%]
        "avg_depth_percent": 14, # средняя глубина [%]
        "depth_mm": 15,          # глубина [мм]
        "remaining_wall": 16,    # остат. ТС [мм]
        "erf_b31g": 20,          # ERF B31G
        "erf_case1": 21,         # ERF (случай 1)
        "erf_case2": 22,         # ERF (случай 2)
        "erf_dnv": 23,           # ERF DNV
        "surface_location": 24,  # локация на поверхн.
        "location_class": 25,    # класс лок.
        "cluster_id": 27,        # № кластера
        "latitude": 28,          # Широта [°]
        "longitude": 29,         # Долгота [°]
        "elevation": 30,         # высота [м]
    }
    
    imported = 0
    errors = []
    
    # Read data rows
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        # Skip empty rows
        if not row[0]:
            continue
        
        try:
            # Generate unique defect code
            defect_code = f"{pipeline_code}-DEF-{uuid.uuid4().hex[:8].upper()}"
            
            # Extract values with safe parsing - ensure all numeric fields are properly converted
            max_depth = parse_float(row[COL_MAP["max_depth_percent"]])
            erf = parse_float(row[COL_MAP["erf_b31g"]])
            
            defect_type = parse_string(row[COL_MAP["defect_type"]])
            if not defect_type:
                continue  # Skip rows without defect type
            
            # Parse all values explicitly to ensure proper types
            weld_dist = parse_float(row[COL_MAP["weld_distance"]])
            section_num = parse_int(row[COL_MAP["section_number"]])
            section_len = parse_float(row[COL_MAP["section_length"]])
            measured_dist = parse_float(row[COL_MAP["measured_distance"]])
            orient = parse_orientation(row[COL_MAP["orientation"]])
            
            length_mm_val = parse_float(row[COL_MAP["length_mm"]])
            width_mm_val = parse_float(row[COL_MAP["width_mm"]])
            avg_depth = parse_float(row[COL_MAP["avg_depth_percent"]])
            depth_mm_val = parse_float(row[COL_MAP["depth_mm"]])
            
            wall_thick = parse_float(row[COL_MAP["wall_thickness"]])
            remain_wall = parse_float(row[COL_MAP["remaining_wall"]])
            
            erf_case1_val = parse_float(row[COL_MAP["erf_case1"]])
            erf_case2_val = parse_float(row[COL_MAP["erf_case2"]])
            erf_dnv_val = parse_float(row[COL_MAP["erf_dnv"]])
            
            lat_val = parse_float(row[COL_MAP["latitude"]])
            lon_val = parse_float(row[COL_MAP["longitude"]])
            elev_val = parse_float(row[COL_MAP["elevation"]])
            
            defect = Defect(
                defect_code=defect_code,
                pipeline_code=pipeline_code,
                
                # Location
                weld_distance=weld_dist,
                section_number=section_num,
                section_length=section_len,
                measured_distance=measured_dist,
                orientation=orient,
                
                # Classification
                defect_type=defect_type,
                identification=parse_string(row[COL_MAP["identification"]]),
                external_size=parse_string(row[COL_MAP["external_size"]]),
                severity=calculate_severity(max_depth, erf),
                
                # Dimensions
                length_mm=length_mm_val,
                width_mm=width_mm_val,
                max_depth_percent=max_depth,
                avg_depth_percent=avg_depth,
                depth_mm=depth_mm_val,
                
                # Wall thickness
                wall_thickness=wall_thick,
                remaining_wall=remain_wall,
                
                # ERF
                erf_b31g=erf,
                erf_case1=erf_case1_val,
                erf_case2=erf_case2_val,
                erf_dnv=erf_dnv_val,
                
                # Location
                surface_location=parse_string(row[COL_MAP["surface_location"]]),
                location_class=parse_string(row[COL_MAP["location_class"]]),
                cluster_id=parse_int(row[COL_MAP["cluster_id"]]),
                
                # Coordinates
                latitude=lat_val,
                longitude=lon_val,
                elevation=elev_val,
                
                # Metadata
                comment=parse_string(row[COL_MAP["comment"]]),
                inspection_date=inspection_date,
            )
            
            db.add(defect)
            imported += 1
            
        except Exception as e:
            errors.append(str(e))
    
    db.commit()
    wb.close()
    
    return {
        "imported": imported,
        "errors": errors[:10] if errors else [],
        "total_errors": len(errors),
        "pipeline_code": pipeline_code,
    }


def get_available_sheets(file_path: str) -> List[str]:
    """Get list of sheet names from Excel file."""
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    return sheets
